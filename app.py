import tkinter as tk
from tkinter import ttk
import openpyxl
import os
from tkinter import messagebox

# CARREGANDO A TABELA
def load_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    else:
        messagebox.showerror ("Erro", "arquivo não encontrado")
        return None 
    
wb_path = "produtos.xlsx"
wb = load_workbook(wb_path)
sheet = wb["Sheet"]
sheet_obj = wb.active
list_values = list(sheet_obj.values)

# CRIANDO AS FUNÇÕES DOS BOTÕES
def insert_row():

    # Pegando as informações dos widgets
    code = code_spinbox.get()
    product = product_entry.get()
    stock_status = status_combobox.get()
    avaliability = "Saiu de linha" if a.get() else "Disponível"

    # Inserindo as informações na tabela .xlsx
    wb_path = "produtos.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    sheet = wb["Sheet"]
    sheet_obj = wb.active
    row_values = [code, product, stock_status, avaliability]
    sheet_obj.append(row_values)
    wb.save(wb_path)

    # Inserindo na visualização da aplicação
    treeview.insert('', tk.END, values=row_values)

    # Resetando as configurações dos widgets
    code_spinbox.delete(0, "end")
    code_spinbox.insert(0, "Código do produto")
    product_entry.delete(0, "end")
    product_entry.insert(0, "Nome do Produto")
    status_combobox.delete(0, "end")
    status_combobox.insert(0, "Status de estoque")
    a.set(False)

# Função para selecionar o produto já cadastrado
def select_product():
    selected_item = treeview.selection()

    if not selected_item:
        messagebox.showwarning("Aviso", "Selecione um produto na lista.")
    else:
        values = treeview.item(selected_item, 'values')
        code_spinbox.delete(0, tk.END)
        code_spinbox.insert(0, values[0])
        product_entry.delete(0, tk.END)
        product_entry.insert(0, values[1])
        status_combobox.delete(0, tk.END)
        status_combobox.insert(0, values[2])
        a.set(True if values[3] == "Saiu de linha" else False)
   
# Função para salvar alterações no produto
def update_product():
    # Pegando as informações dos widgets
    selected_item = treeview.selection()

    if not selected_item:
        messagebox.showwarning("Aviso", "Selecione um produto na lista para editar.")
        return

    code = code_spinbox.get()
    product = product_entry.get()
    stock_status = status_combobox.get()
    avaliability = "Saiu de linha" if a.get() else "Disponível"

    # Atualizando as informações na tabela .xlsx
    wb_path = "produtos.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    sheet = wb["Sheet"]
    sheet_obj = wb.active

    # Encontrando a linha correspondente ao código do produto selecionado
    for row in sheet_obj.iter_rows(min_row=2):
        if str(row[0].value).strip() == str(code).strip():  # Convertendo para string antes de usar strip()
            # Atualizando os valores da linha encontrada
            row_values = [code, product, stock_status, avaliability]
            for cell, value in zip(row, row_values):
                cell.value = value
            wb.save(wb_path)
            treeview.item(selected_item, values=row_values)
            messagebox.showinfo("Sucesso", "Produto atualizado com sucesso.")
            return

    messagebox.showerror("Erro", "Produto não encontrado na planilha.")
    
# Função para deletar o produto
def delete_product():
    selected_item = treeview.selection()

    if not selected_item:
        messagebox.showwarning("Aviso", "Selecione um produto na lista para deletar.")
    else:
        wb_path = "produtos.xlsx"
        wb = openpyxl.load_workbook(wb_path)
        sheet = wb["Sheet"]
        sheet_obj = wb.active
        code = treeview.item(selected_item, 'values')[0]

        for row in sheet_obj.iter_rows(min_row=2):
            if str(row[0].value).strip() == str(code).strip():  # Convertendo para string antes de usar strip()
                idx = row[0].row
                sheet_obj.delete_rows(idx)
                wb.save(wb_path)
                treeview.delete(selected_item)
                messagebox.showinfo("Sucesso", "Produto deletado com sucesso.")
                return

        messagebox.showerror("Erro", "Produto não encontrado na planilha.")

# CONFIGURANDO A INTERFACE GRÁFICA - GUI
root= tk.Tk()
root.title("Sistema de Cadastro de Produtos")

frame = tk.Frame(root)
frame.pack()

# Moldura dos campos de preenchimento
widgets_frame = ttk.LabelFrame(frame, text="Dados do Produto")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

# Campo do código
code_spinbox = ttk.Spinbox(widgets_frame, from_="0000", to="9999")
code_spinbox.insert(0, "Código do produto")
code_spinbox.bind("<FocusIn>", lambda e: product_entry.delete(0, 'end') )
code_spinbox.grid(row=0 , column=0, padx=5, pady=5, sticky="ew")

# Campo do produto
product_entry = ttk.Entry(widgets_frame)
product_entry.insert(0, "Nome do Produto")
product_entry.bind("<FocusIn>", lambda e: product_entry.delete(0, 'end') )
product_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

# Status do estoque
combo_list = ("Possui estoque", "Em falta")
status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
status_combobox.insert(0, "Status de estoque")
status_combobox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

# Disponibilidade do produto
a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widgets_frame, text="Saiu de linha", variable=a)
checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

# Botão para inserir
insert_button = ttk.Button(widgets_frame, text="Salvar produto", command= insert_row)
insert_button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

# Separar campos 1
separator = ttk.Separator(widgets_frame)
separator.grid(row=5, column=0, padx=10, pady=10, sticky="ew")

# Botão para selecionar produto
select_button = ttk.Button(widgets_frame, text="Selecionar produto já cadastrado", command= select_product)
select_button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

# Botão para editar produto
update_button = ttk.Button(widgets_frame, text="Salvar alterações do produto", command= update_product)
update_button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")

# Separar campos 2
separator = ttk.Separator(widgets_frame)
separator.grid(row=8, column=0, padx=10, pady=10, sticky="ew")

# Botão para deletar produto
delete_button = ttk.Button(widgets_frame, text="Deletar produto", command= delete_product)
delete_button.grid(row=9, column=0, padx=5, pady=5, sticky="nsew")

# Visualização da tabela Excel
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

view_xl = ("CÓDIGO", "PRODUTO", "STATUS", "DISPONIBILIDADE")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=view_xl, height=13)
treeview.column("CÓDIGO", width=80)
treeview.column("PRODUTO", width=150)
treeview.column("STATUS", width=150)
treeview.column("DISPONIBILIDADE", width=150)
treeview.pack()
treeScroll.config(command=treeview.yview)

# Inserindo os valores na janela da aplicação
for col_name in list_values[0]:
    treeview.heading(col_name, text=col_name)
for value_tuple in list_values[1:]:
    treeview.insert('', tk.END, values=value_tuple)



root.mainloop()